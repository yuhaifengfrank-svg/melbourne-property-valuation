"""
ingest_rba_macro_full.py — RBA macro indicators (CPI, GDP, Labour, Cash Rate) → summary artifact

Downloads (expected):
  f01d.xlsx    — Cash Rate Target                       (daily, 2011→)
  g01hist.xlsx — Consumer Price Inflation (actual CPI)   (quarterly, 1922→)
  h01hist.xlsx — GDP (actual)                            (quarterly, 1959→)
  h05hist.xlsx — Labour Force (unemployment, etc.)       (monthly, 1978→)
  j1-*.csv     — Market forecasts (GDP/CPI/unemployment) (quarterly)

Output:
  data/artifacts/macro_summary/rba_macro_full.json
"""

import json
import os
import openpyxl
import csv
from datetime import datetime

PROJECT_ROOT = "/opt/aushomevalue"
RAW_DIR = os.path.join(PROJECT_ROOT, "data", "raw", "rba")
ARTIFACTS_DIR = os.path.join(PROJECT_ROOT, "data", "artifacts", "macro_summary")
os.makedirs(ARTIFACTS_DIR, exist_ok=True)


def parse_date(val):
    """Parse datetime or date string to 'YYYY-MM-DD'."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        try:
            return datetime.strptime(val.strip()[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            return val.strip()[:10]
    return None


def parse_float(val):
    """Parse to float or None."""
    if val is None or val == "" or val == "-":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def read_xlsx_series(path, sheet, data_start_row, columns):
    """
    Read an RBA xlsx with metadata rows, extract named columns.
    columns: list of (col_idx, name) — col_idx 0-based
    Returns: list of {date, name1: val1, name2: val2, ...}
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    series_row = rows[data_start_row - 1] if data_start_row > 0 else None

    records = []
    for row in rows[data_start_row:]:
        date_val = row[0]
        if not date_val or str(date_val).strip() in ('', '-'):
            continue
        date_str = parse_date(date_val)
        if not date_str:
            continue

        rec = {"date": date_str}
        for col_idx, name in columns:
            rec[name] = parse_float(row[col_idx])
        # Only keep if at least one column has a value
        if not any(rec[n] is not None for _, n in columns):
            continue
        records.append(rec)

    wb.close()
    return records


def read_j1_csv(path):
    """Read RBA J1 CSV (8 header rows, data from row 9)."""
    records = []
    with open(path, newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i < 8:
                continue
            if not row or not row[0].strip():
                continue
            # Row format: [date, target_quarter, median, mean, low, high, responses]
            rec = {
                "survey_date": row[0].strip() if len(row) > 0 else None,
                "target_quarter": row[1].strip() if len(row) > 1 else None,
                "median": parse_float(row[2]) if len(row) > 2 else None,
                "mean": parse_float(row[3]) if len(row) > 3 else None,
                "low": parse_float(row[4]) if len(row) > 4 else None,
                "high": parse_float(row[5]) if len(row) > 5 else None,
            }
            records.append(rec)
    return records


def main():
    print("=== RBA Full Macro → Summary ===\n")

    # ── 1) Cash Rate (f01d.xlsx) ──
    print("[1/4] Cash Rate Target...")
    path = os.path.join(RAW_DIR, "f01d.xlsx")
    cash_rate_raw = read_xlsx_series(path, "Data", 10, [(1, "cash_rate_target")])
    # Extract latest + changes
    latest_rate = None
    latest_date = None
    changes = []
    full_cash = []
    for r in cash_rate_raw:
        if r["cash_rate_target"] is not None:
            full_cash.append(r)
            latest_rate = r["cash_rate_target"]
            latest_date = r["date"]

    # Change tracking from Column C (index 2) — need second pass
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Data"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[10:]:
        if not row[0]:
            continue
        ds = parse_date(row[0])
        chg = parse_float(row[2])
        if chg is not None and ds:
            changes.append({"date": ds, "change": chg})
    wb.close()
    changes = changes[-10:][::-1]

    print(f"   {len(full_cash)} daily records, latest: {latest_date} @ {latest_rate}%")

    # ── 2) CPI — G1 (actual quarterly) ──
    print("[2/4] CPI (Consumer Price Inflation)...")
    path = os.path.join(RAW_DIR, "g01hist.xlsx")
    cpi_raw = read_xlsx_series(path, "Data", 10, [
        (1, "cpi_index"),
        (2, "cpi_yoy"),
        (3, "cpi_ex_volatile_yoy"),
    ])
    cpi_latest = cpi_raw[-1] if cpi_raw else None
    print(f"   {len(cpi_raw)} quarterly records, latest CPI YoY: {cpi_latest.get('cpi_yoy') if cpi_latest else 'N/A'}%")

    # ── 3) GDP — H1 (actual quarterly) ──
    print("[3/4] GDP (Gross Domestic Product)...")
    path = os.path.join(RAW_DIR, "h01hist.xlsx")
    gdp_raw = read_xlsx_series(path, "Data", 10, [
        (1, "real_gdp_audm"),
        (2, "real_gdp_yoy"),
    ])
    gdp_latest = gdp_raw[-1] if gdp_raw else None
    print(f"   {len(gdp_raw)} quarterly records, latest GDP YoY: {gdp_latest.get('real_gdp_yoy') if gdp_latest else 'N/A'}%")

    # ── 4) Labour Force — H5 (monthly) ──
    print("[4/4] Labour Force (Unemployment)...")
    path = os.path.join(RAW_DIR, "h05hist.xlsx")
    labour_raw = read_xlsx_series(path, "Data", 10, [
        (1, "labour_force_000"),
        (2, "participation_rate"),
        (5, "employment_000"),
        (9, "unemployment_000"),
        (10, "unemployment_rate_sa"),
        (11, "unemployment_rate_trend"),
    ])
    labour_latest = labour_raw[-1] if labour_raw else None
    print(f"   {len(labour_raw)} monthly records, latest UE rate: {labour_latest.get('unemployment_rate_sa') if labour_latest else 'N/A'}%")

    # ── 5) J1 Forecasts (CSV) ──
    print("   Reading J1 forecasts (GDP, CPI, Unemployment)...")
    j1_sources = {
        "gdp_growth": "j1-gdp-growth.csv",
        "headline_inflation": "j1-headline-inflation.csv",
        "underlying_inflation": "j1-underlying-inflation.csv",
        "unemployment_rate": "j1-unemployment-rate.csv",
    }
    forecasts = {}
    for name, fname in j1_sources.items():
        fpath = os.path.join(RAW_DIR, fname)
        if os.path.exists(fpath):
            recs = read_j1_csv(fpath)
            forecasts[name] = {
                "total_records": len(recs),
                "latest": recs[-1] if recs else None,
            }
            print(f"     J1 {name}: {len(recs)} records")

    # ── Build artifact ──
    artifact = {
        "source": "RBA Macro Indicators (F1 + G1 + H1 + H5 + J1)",
        "extracted_at": "2026-06-23",
        "series": {
            "cash_rate_target": {
                "frequency": "daily",
                "period": f"{full_cash[0]['date']} → {full_cash[-1]['date']}" if full_cash else "N/A",
                "total_records": len(full_cash),
                "latest": {"date": latest_date, "value": latest_rate},
                "recent_changes": changes,
                "records": full_cash[-120:]  # last ~4 months to keep size manageable
            },
            "consumer_price_index": {
                "frequency": "quarterly",
                "source": "ABS / RBA (G1)",
                "total_records": len(cpi_raw),
                "latest": cpi_latest,
                "records": cpi_raw[-80:]  # last 20 years
            },
            "gdp": {
                "frequency": "quarterly",
                "source": "ABS / RBA (H1)",
                "total_records": len(gdp_raw),
                "latest": gdp_latest,
                "records": gdp_raw[-80:]
            },
            "labour_force": {
                "frequency": "monthly",
                "source": "ABS / RBA (H5)",
                "total_records": len(labour_raw),
                "latest": labour_latest,
                "records": labour_raw[-120:]
            },
            "forecasts": {
                "source": "RBA J1 Market Economists' Survey",
                "latest_publication": "08-May-2026",
                "series": forecasts
            }
        }
    }

    output_path = os.path.join(ARTIFACTS_DIR, "rba_macro_full.json")
    with open(output_path, "w") as f:
        json.dump(artifact, f, indent=2)

    file_size = os.path.getsize(output_path)
    print(f"\n✅ Written: {output_path}")
    print(f"   File size: {file_size:,} bytes")
    print(f"   Contains: Cash Rate, CPI, GDP, Labour Force + J1 forecasts")


if __name__ == "__main__":
    main()
