"""
ingest_abs_census.py — ABS Census 2021 SA2 → suburb summary artifact

Flow:
  1. Read raw CSV from data/raw/abs/2021 Census GCP Statistical Area 2 for VIC/
  2. Parse key tables: G01 (pop), G02 (median age/income/rent), G06 (housing costs), G41 (dwelling structure)
  3. Aggregate SA2 → suburb/LGA mapping using Metadata/2021Census_geog_desc_1st...csv
  4. Output suburb-level summary JSON → data/artifacts/suburb_summary/

Usage:
  python3 app/ingest/ingest_abs_census.py

Output:
  data/artifacts/suburb_summary/abs_census_2021.json
"""

import csv
import json
import os
from collections import defaultdict

PROJECT_ROOT = "/opt/aushomevalue"
RAW_DIR = os.path.join(PROJECT_ROOT, "data", "raw", "abs", "2021 Census GCP Statistical Area 2 for VIC")
ARTIFACTS_DIR = os.path.join(PROJECT_ROOT, "data", "artifacts", "suburb_summary")
os.makedirs(ARTIFACTS_DIR, exist_ok=True)

def read_csv(filename):
    """Read a CSV from RAW_DIR, return list of dicts."""
    path = os.path.join(RAW_DIR, filename)
    if not os.path.exists(path):
        print(f"[WARN] File not found: {path}")
        return []
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        return list(reader)

def parse_int(val):
    try:
        return int(val) if val not in (None, "", "NA", "NFP") else 0
    except (ValueError, TypeError):
        return 0

def parse_float(val):
    try:
        return float(val) if val not in (None, "", "NA", "NFP") else None
    except (ValueError, TypeError):
        return None

def parse_geog_desc():
    """Read SA2→name mapping from ABS geography xlsx (ASGS Non-ABS Structures sheet)."""
    geog_dir = os.path.join(PROJECT_ROOT, "data", "raw", "abs", "Metadata")
    geog_file = None
    for f in os.listdir(geog_dir):
        if "geog_desc" in f.lower():
            geog_file = os.path.join(geog_dir, f)
            break
    
    if not geog_file:
        print("[WARN] Geography description file not found")
        return {}
    
    sa2_to_name = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(geog_file, read_only=True)
        # Look for SA2 sheet or Non-ABS Structures sheet
        target_sheet = None
        for sn in wb.sheetnames:
            if "SA2" in sn or "Non_ABS" in sn:
                target_sheet = sn
                break
        if not target_sheet:
            print(f"[WARN] No SA2/Non_ABS sheet found in {geog_file}, sheets: {wb.sheetnames[:5]}...")
            return {}
        ws = wb[target_sheet]
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            asgs_type = str(row[0]).strip()
            code = str(row[1]).strip() if len(row) > 1 else ""
            name = str(row[3]).strip() if len(row) > 3 else ""
            # SA2 codes are numeric, 9 digits — filter for SA2 level
            if asgs_type == "SA2" and code.isdigit():
                sa2_to_name[code] = name
        print(f"  SA2 codes mapped from ASGS: {len(sa2_to_name)}")
    except ImportError:
        print("[WARN] openpyxl not installed")
        return {}
    
    return sa2_to_name

def main():
    print("=== ABS Census 2021 → Suburb Summary ===")
    
    # 1. Geography mapping
    print("[1/4] Reading geography mapping...")
    sa2_to_name = parse_geog_desc()
    
    # 2. G01 — Population by age/gender
    print("[2/4] Reading G01 (population)...")
    g01 = read_csv("2021Census_G01_VIC_SA2.csv")
    
    sa2_data = {}
    for row in g01:
        code = row.get("SA2_CODE_2021", "").strip()
        if not code:
            continue
        
        sa2_data[code] = {
            "sa2_code": code,
            "sa2_name": sa2_to_name.get(code, ""),
            "total_population": parse_int(row.get("Tot_P_P", "0")),
            "total_male": parse_int(row.get("Tot_P_M", "0")),
            "total_female": parse_int(row.get("Tot_P_F", "0")),
        }
    
    print(f"  SA2 records: {len(sa2_data)}")
    
    # 3. G02 — Median age / income / rent / mortgage
    print("[3/4] Reading G02 (median indicators)...")
    g02 = read_csv("2021Census_G02_VIC_SA2.csv")
    
    for row in g02:
        code = row.get("SA2_CODE_2021", "").strip()
        if not code or code not in sa2_data:
            continue
        
        sa2_data[code].update({
            "median_age": parse_float(row.get("Median_age_persons")),
            "median_mortgage_repay": parse_float(row.get("Median_mortgage_repay_monthly")),
            "median_rent_weekly": parse_float(row.get("Median_rent_weekly")),
            "median_personal_income_weekly": parse_float(row.get("Median_tot_prsnl_inc_weekly")),
            "median_family_income_weekly": parse_float(row.get("Median_tot_fam_inc_weekly")),
            "median_household_income_weekly": parse_float(row.get("Median_tot_hhd_inc_weekly")),
        })
    
    # 4. G41 — Dwelling structure (bedrooms by type)
    print("[4/4] Reading G41 (dwelling structure)...")
    g41 = read_csv("2021Census_G41_VIC_SA2.csv")
    
    for row in g41:
        code = row.get("SA2_CODE_2021", "").strip()
        if not code or code not in sa2_data:
            continue
        
        # Total dwellings (last column: Total_Total)
        keys = list(row.keys())
        total_dwellings = parse_int(row.get(keys[-1], "0")) if keys else 0
        
        # Separate houses total
        sep_house_total = parse_int(row.get("Separate_house_Total", "0"))
        
        # Total bedrooms (various columns)
        bd_0 = parse_int(row.get("Total_bedrooms_Bedroom_0_count", "0"))
        bd_1 = parse_int(row.get("Total_bedrooms_Bedroom_1_count", "0"))
        bd_2 = parse_int(row.get("Total_bedrooms_Bedroom_2_count", "0"))
        bd_3 = parse_int(row.get("Total_bedrooms_Bedroom_3_count", "0"))
        bd_4 = parse_int(row.get("Total_bedrooms_Bedroom_4_count", "0"))
        
        sa2_data[code].update({
            "total_dwellings": total_dwellings,
            "separate_houses": sep_house_total,
            "bedrooms_0_1": bd_0 + bd_1,
            "bedrooms_2": bd_2,
            "bedrooms_3": bd_3,
            "bedrooms_4_plus": bd_4,
        })
    
    # 5. Write artifact
    output_path = os.path.join(ARTIFACTS_DIR, "abs_census_2021.json")
    records = sorted(sa2_data.values(), key=lambda x: x["sa2_code"])
    
    with open(output_path, "w") as f:
        json.dump({
            "source": "ABS Census 2021 GCP SA2 (VIC)",
            "url": "https://www.abs.gov.au/census/find-census-data/datapacks",
            "downloaded_at": "2026-06-23",
            "total_sa2": len(records),
            "records": records
        }, f, indent=2)
    
    print(f"\n✅ Written: {output_path}")
    print(f"   Total SA2 records: {len(records)}")
    print(f"   File size: {os.path.getsize(output_path):,} bytes")

if __name__ == "__main__":
    main()
